"""
small_world_agents.py – CRUD, AI generation, bulk import, and relationship
management for Small World agents.
"""

from __future__ import annotations

import io
import json
import uuid as _uuid
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models, schemas
from ..deps import get_current_user, get_db, require_credits

router = APIRouter(prefix="/api/small-world", tags=["small-world-agents"])


# ── Helpers ───────────────────────────────────────────────────────────────────

def _get_world_or_404(world_id: str, user_id: int, db: Session) -> models.SmallWorld:
    world = db.query(models.SmallWorld).filter(
        models.SmallWorld.world_id == world_id,
        models.SmallWorld.user_id == user_id,
    ).first()
    if not world:
        raise HTTPException(status_code=404, detail="World not found")
    return world


def _agent_to_response(agent: models.SmallWorldAgent, db: Session, world_uuid: str | None = None) -> schemas.AgentResponse:
    rel_count = db.query(models.AgentRelationship).filter(
        (models.AgentRelationship.source_agent_id == agent.id) |
        (models.AgentRelationship.target_agent_id == agent.id)
    ).count()

    def _load(field: str | None) -> dict | None:
        if field is None:
            return None
        try:
            return json.loads(field)
        except Exception:
            return None

    return schemas.AgentResponse(
        id=agent.id,
        agent_id=agent.agent_id,
        world_id=world_uuid,
        name=agent.name,
        age=agent.age,
        gender=agent.gender,
        location=agent.location,
        profession=agent.profession,
        job_title=agent.job_title,
        organization=agent.organization,
        personality_traits=_load(agent.personality_traits),
        behavioral_attributes=_load(agent.behavioral_attributes),
        contextual_state=_load(agent.contextual_state),
        external_factors=_load(agent.external_factors),
        created_at=agent.created_at,
        updated_at=agent.updated_at,
        relationship_count=rel_count,
    )


def _relationship_to_response(
    relationship: models.AgentRelationship,
    agent_lookup: dict[int, models.SmallWorldAgent],
) -> schemas.AgentRelationshipResponse:
    source = agent_lookup.get(relationship.source_agent_id)
    target = agent_lookup.get(relationship.target_agent_id)
    return schemas.AgentRelationshipResponse(
        id=relationship.id,
        rel_id=relationship.rel_id,
        source_agent_id=source.agent_id if source else "",
        target_agent_id=target.agent_id if target else "",
        source_agent_name=source.name if source else None,
        target_agent_name=target.name if target else None,
        type=relationship.type,
        strength=relationship.strength,
        sentiment=relationship.sentiment,
        influence_direction=relationship.influence_direction,
    )


def _apply_agent_data(agent: models.SmallWorldAgent, data: schemas.AgentCreate | schemas.AgentUpdate) -> None:
    """Apply create/update payload onto an ORM object."""
    simple_fields = ["name", "age", "gender", "location", "profession", "job_title", "organization"]
    for f in simple_fields:
        v = getattr(data, f, None)
        if v is not None:
            setattr(agent, f, v)

    json_fields = ["personality_traits", "behavioral_attributes", "contextual_state", "external_factors"]
    for f in json_fields:
        v = getattr(data, f, None)
        if v is not None:
            setattr(agent, f, json.dumps(v.model_dump(exclude_none=False)))


# ── List agents ───────────────────────────────────────────────────────────────

@router.get("/worlds/{world_id}/agents/", response_model=list[schemas.AgentResponse])
def list_agents(
    world_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agents = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.world_id == world.id
    ).order_by(models.SmallWorldAgent.created_at.desc()).all()
    return [_agent_to_response(a, db, world_uuid=world.world_id) for a in agents]


# ── Create agent ──────────────────────────────────────────────────────────────

@router.post("/worlds/{world_id}/agents/", response_model=schemas.AgentResponse, status_code=201)
def create_agent(
    world_id: str,
    body: schemas.AgentCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = models.SmallWorldAgent(user_id=current_user.id, world_id=world.id, name=body.name)
    _apply_agent_data(agent, body)
    db.add(agent)
    db.commit()
    db.refresh(agent)
    return _agent_to_response(agent, db, world_uuid=world.world_id)


# ── Download Template ─────────────────────────────────────────────────────────

@router.get("/worlds/{world_id}/agents/template")
def download_template(world_id: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agents"

    # Required / scalar columns
    scalar_columns = [
        "name", "age", "gender", "location", "profession", "job_title", "organization",
        "openness", "conscientiousness", "extraversion", "agreeableness", "neuroticism",
        "risk_tolerance", "decision_style", "core_beliefs",
        "communication_style", "influence_level", "adaptability", "loyalty", "stress_response",
        "salary", "work_environment", "market_exposure",
    ]
    # Optional list columns – values should be semicolon-separated (e.g. "goal1;goal2")
    list_columns = [
        "motivation_drivers", "biases",
        "current_goals", "current_frustrations", "incentives", "constraints",
    ]
    columns = scalar_columns + list_columns

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    optional_fill = PatternFill(start_color="2d5a8e", end_color="2d5a8e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, col_name in enumerate(columns, start=1):
        is_optional = col_name in list_columns
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = optional_fill if is_optional else header_fill
        cell.alignment = header_align
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 18)

    ws.row_dimensions[1].height = 32

    # Example row – list fields use semicolons as separator
    example = [
        "Jane Smith", 34, "Female", "New York, USA", "Product Management", "Senior PM", "Acme Corp",
        0.7, 0.8, 0.6, 0.7, 0.4,
        0.5, "analytical", "Innovation drives progress",
        "direct", 0.7, 0.8, 0.6, "Seeks clarity",
        "$120,000", "Hybrid remote", "SaaS B2B",
        # list columns
        "achievement;impact",   # motivation_drivers
        "optimism bias",        # biases
        "Launch v2;Grow team",  # current_goals
        "slow feedback loops",  # current_frustrations
        "equity;bonus",         # incentives
        "budget;headcount",     # constraints
    ]
    for col_idx, val in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    # Row 3: instructions for list columns
    note_row = 3
    for col_idx, col_name in enumerate(columns, start=1):
        if col_name in list_columns:
            ws.cell(row=note_row, column=col_idx, value="(separate multiple values with ;)")
            ws.cell(row=note_row, column=col_idx).font = Font(italic=True, color="888888", size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=small_world_agents_template.xlsx"},
    )


# ── Get agent graph (must come before /{agent_id}) ───────────────────────────

@router.get("/worlds/{world_id}/agents/graph", response_model=schemas.AgentGraphResponse)
def get_agent_graph(
    world_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agents = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.world_id == world.id
    ).order_by(models.SmallWorldAgent.created_at.desc()).all()

    if not agents:
        return schemas.AgentGraphResponse(agents=[], relationships=[])

    agent_lookup = {agent.id: agent for agent in agents}
    world_agent_ids = list(agent_lookup.keys())
    relationships = db.query(models.AgentRelationship).filter(
        models.AgentRelationship.source_agent_id.in_(world_agent_ids),
        models.AgentRelationship.target_agent_id.in_(world_agent_ids),
    ).all()

    return schemas.AgentGraphResponse(
        agents=[_agent_to_response(agent, db, world_uuid=world_id) for agent in agents],
        relationships=[
            _relationship_to_response(relationship, agent_lookup)
            for relationship in relationships
        ],
    )


# ── Get single agent ──────────────────────────────────────────────────────────

@router.get("/worlds/{world_id}/agents/{agent_id}", response_model=schemas.AgentResponse)
def get_agent(
    world_id: str,
    agent_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.agent_id == agent_id,
        models.SmallWorldAgent.world_id == world.id,
    ).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    return _agent_to_response(agent, db, world_uuid=world.world_id)


# ── Update agent ──────────────────────────────────────────────────────────────

@router.put("/worlds/{world_id}/agents/{agent_id}", response_model=schemas.AgentResponse)
def update_agent(
    world_id: str,
    agent_id: str,
    body: schemas.AgentUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.agent_id == agent_id,
        models.SmallWorldAgent.world_id == world.id,
    ).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    _apply_agent_data(agent, body)
    db.commit()
    db.refresh(agent)
    return _agent_to_response(agent, db, world_uuid=world.world_id)


# ── Delete agent ──────────────────────────────────────────────────────────────

@router.delete("/worlds/{world_id}/agents/{agent_id}", status_code=204)
def delete_agent(
    world_id: str,
    agent_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.agent_id == agent_id,
        models.SmallWorldAgent.world_id == world.id,
    ).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    db.delete(agent)
    db.commit()


# ── AI Agent Generation ───────────────────────────────────────────────────────

@router.post("/worlds/{world_id}/agents/generate", response_model=schemas.AgentCreate)
def generate_agent(
    world_id: str,
    body: schemas.AgentGenerateRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credits),
):
    _get_world_or_404(world_id, current_user.id, db)
    """
    Take sparse fields + natural language description and use Gemini to produce
    a complete AgentCreate payload. The frontend shows this for user review
    before calling POST /agents/ to save.
    """
    from core.agent_generator import generate_agent_profile
    from ..billing import deduct_credits

    sparse = {
        "name": body.name,
        "profession": body.profession,
        "organization": body.organization,
        "location": body.location,
        "age": body.age,
        "description": body.description,
    }
    try:
        profile, usage = generate_agent_profile(sparse)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"AI generation failed: {exc}")

    try:
        deduct_credits(db, current_user.id, usage, description="Agent AI generation")
    except Exception as billing_exc:
        print(f"[billing] deduct_credits failed: {billing_exc}")

    # Validate and return
    try:
        return schemas.AgentCreate(**profile)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"AI returned invalid profile: {exc}")


# ── Bulk Import (Excel / CSV) ─────────────────────────────────────────────────

@router.post("/worlds/{world_id}/agents/bulk-import", response_model=list[schemas.AgentResponse], status_code=201)
async def bulk_import_agents(
    world_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    import pandas as pd

    content = await file.read()
    buf = io.BytesIO(content)

    filename = (file.filename or "").lower()
    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(buf)
        else:
            df = pd.read_excel(buf)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse file: {exc}")

    required_column = "name"
    if required_column not in df.columns:
        raise HTTPException(status_code=422, detail="File must contain a 'name' column")

    created = []
    for _, row in df.iterrows():
        name = str(row.get("name", "")).strip()
        if not name:
            continue

        def _safe(col: str) -> str | None:
            v = row.get(col)
            if v is None or (isinstance(v, float) and str(v) == "nan"):
                return None
            return str(v).strip() or None

        def _safe_int(col: str) -> int | None:
            v = row.get(col)
            try:
                return int(v)
            except Exception:
                return None

        def _safe_float(col: str) -> float | None:
            v = row.get(col)
            try:
                return float(v)
            except Exception:
                return None

        def _safe_list(col: str) -> list[str] | None:
            """Parse a semicolon-separated string or JSON array into a list."""
            v = row.get(col)
            if v is None or (isinstance(v, float) and str(v) == "nan"):
                return None
            raw = str(v).strip()
            if not raw:
                return None
            # Try JSON array first
            if raw.startswith("["):
                try:
                    parsed = json.loads(raw)
                    if isinstance(parsed, list):
                        return [str(i).strip() for i in parsed if str(i).strip()]
                except Exception:
                    pass
            # Fall back to semicolon-separated
            items = [s.strip() for s in raw.split(";") if s.strip()]
            return items if items else None

        personality_traits = schemas.PersonalityTraits(
            openness=_safe_float("openness"),
            conscientiousness=_safe_float("conscientiousness"),
            extraversion=_safe_float("extraversion"),
            agreeableness=_safe_float("agreeableness"),
            neuroticism=_safe_float("neuroticism"),
            risk_tolerance=_safe_float("risk_tolerance"),
            decision_style=_safe("decision_style"),
            motivation_drivers=_safe_list("motivation_drivers"),
            core_beliefs=_safe("core_beliefs"),
            biases=_safe_list("biases"),
        )
        behavioral_attributes = schemas.BehavioralAttributes(
            communication_style=_safe("communication_style"),
            influence_level=_safe_float("influence_level"),
            adaptability=_safe_float("adaptability"),
            loyalty=_safe_float("loyalty"),
            stress_response=_safe("stress_response"),
        )
        contextual_state = schemas.ContextualState(
            current_goals=_safe_list("current_goals"),
            current_frustrations=_safe_list("current_frustrations"),
            incentives=_safe_list("incentives"),
            constraints=_safe_list("constraints"),
        )
        external_factors = schemas.ExternalFactors(
            salary=_safe("salary"),
            work_environment=_safe("work_environment"),
            market_exposure=_safe("market_exposure"),
        )

        agent = models.SmallWorldAgent(
            user_id=current_user.id,
            world_id=world.id,
            name=name,
            age=_safe_int("age"),
            gender=_safe("gender"),
            location=_safe("location"),
            profession=_safe("profession"),
            job_title=_safe("job_title"),
            organization=_safe("organization"),
            personality_traits=json.dumps(personality_traits.model_dump()),
            behavioral_attributes=json.dumps(behavioral_attributes.model_dump()),
            contextual_state=json.dumps(contextual_state.model_dump()),
            external_factors=json.dumps(external_factors.model_dump()),
        )
        db.add(agent)
        db.flush()
        created.append(agent)

    db.commit()
    for a in created:
        db.refresh(a)
    return [_agent_to_response(a, db, world_uuid=world.world_id) for a in created]


# ── Relationships ─────────────────────────────────────────────────────────────

def _get_agent_or_404(agent_id: str, world_db_id: int, db: Session) -> models.SmallWorldAgent:
    agent = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.agent_id == agent_id,
        models.SmallWorldAgent.world_id == world_db_id,
    ).first()
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    return agent


@router.get("/worlds/{world_id}/agents/{agent_id}/relationships", response_model=list[schemas.AgentRelationshipResponse])
def list_relationships(
    world_id: str,
    agent_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = _get_agent_or_404(agent_id, world.id, db)
    rels = db.query(models.AgentRelationship).filter(
        (models.AgentRelationship.source_agent_id == agent.id) |
        (models.AgentRelationship.target_agent_id == agent.id)
    ).all()

    related_agent_ids = {
        relationship.source_agent_id for relationship in rels
    } | {
        relationship.target_agent_id for relationship in rels
    }
    agent_lookup = {
        related_agent.id: related_agent
        for related_agent in db.query(models.SmallWorldAgent).filter(
            models.SmallWorldAgent.id.in_(related_agent_ids)
        ).all()
    }

    return [_relationship_to_response(relationship, agent_lookup) for relationship in rels]


@router.post("/worlds/{world_id}/agents/{agent_id}/relationships", response_model=schemas.AgentRelationshipResponse, status_code=201)
def create_relationship(
    world_id: str,
    agent_id: str,
    body: schemas.AgentRelationshipCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    source = _get_agent_or_404(agent_id, world.id, db)
    target = _get_agent_or_404(body.target_agent_id, world.id, db)

    if source.id == target.id:
        raise HTTPException(status_code=400, detail="Cannot create self-relationship")

    # Check duplicate
    existing = db.query(models.AgentRelationship).filter(
        models.AgentRelationship.source_agent_id == source.id,
        models.AgentRelationship.target_agent_id == target.id,
        models.AgentRelationship.type == body.type,
    ).first()
    if existing:
        raise HTTPException(status_code=409, detail="Relationship already exists")

    rel = models.AgentRelationship(
        source_agent_id=source.id,
        target_agent_id=target.id,
        type=body.type,
        strength=body.strength or 0.5,
        sentiment=body.sentiment or "neutral",
        influence_direction=body.influence_direction or "both",
    )
    db.add(rel)
    db.commit()
    db.refresh(rel)

    return schemas.AgentRelationshipResponse(
        id=rel.id,
        rel_id=rel.rel_id,
        source_agent_id=source.agent_id,
        target_agent_id=target.agent_id,
        source_agent_name=source.name,
        target_agent_name=target.name,
        type=rel.type,
        strength=rel.strength,
        sentiment=rel.sentiment,
        influence_direction=rel.influence_direction,
    )


@router.delete("/worlds/{world_id}/agents/{agent_id}/relationships/{rel_id}", status_code=204)
def delete_relationship(
    world_id: str,
    agent_id: str,
    rel_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = _get_agent_or_404(agent_id, world.id, db)
    rel = db.query(models.AgentRelationship).filter(
        models.AgentRelationship.rel_id == rel_id,
        (
            (models.AgentRelationship.source_agent_id == agent.id) |
            (models.AgentRelationship.target_agent_id == agent.id)
        ),
    ).first()
    if not rel:
        raise HTTPException(status_code=404, detail="Relationship not found")
    db.delete(rel)
    db.commit()


@router.patch("/worlds/{world_id}/agents/{agent_id}/relationships/{rel_id}", response_model=schemas.AgentRelationshipResponse)
def update_relationship(
    world_id: str,
    agent_id: str,
    rel_id: str,
    body: schemas.AgentRelationshipUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    world = _get_world_or_404(world_id, current_user.id, db)
    agent = _get_agent_or_404(agent_id, world.id, db)
    rel = db.query(models.AgentRelationship).filter(
        models.AgentRelationship.rel_id == rel_id,
        (
            (models.AgentRelationship.source_agent_id == agent.id) |
            (models.AgentRelationship.target_agent_id == agent.id)
        ),
    ).first()
    if not rel:
        raise HTTPException(status_code=404, detail="Relationship not found")

    if body.type is not None and body.type != rel.type:
        conflict = db.query(models.AgentRelationship).filter(
            models.AgentRelationship.source_agent_id == rel.source_agent_id,
            models.AgentRelationship.target_agent_id == rel.target_agent_id,
            models.AgentRelationship.type == body.type,
            models.AgentRelationship.id != rel.id,
        ).first()
        if conflict:
            raise HTTPException(status_code=409, detail="Relationship of this type already exists")
        rel.type = body.type
    if body.strength is not None:
        rel.strength = body.strength
    if body.sentiment is not None:
        rel.sentiment = body.sentiment
    if body.influence_direction is not None:
        rel.influence_direction = body.influence_direction

    db.commit()
    db.refresh(rel)

    agent_lookup = {
        a.id: a for a in db.query(models.SmallWorldAgent).filter(
            models.SmallWorldAgent.id.in_([rel.source_agent_id, rel.target_agent_id])
        ).all()
    }
    return _relationship_to_response(rel, agent_lookup)


# ── Auto-suggest relationships ────────────────────────────────────────────────

@router.post("/worlds/{world_id}/agents/auto-suggest-relationships", response_model=list[schemas.AgentRelationshipResponse])
def auto_suggest_relationships(
    world_id: str,
    body: schemas.AutoSuggestRelationshipsRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_credits),
):
    from core.agent_generator import suggest_relationships
    from ..billing import deduct_credits

    world = _get_world_or_404(world_id, current_user.id, db)
    agents_db = db.query(models.SmallWorldAgent).filter(
        models.SmallWorldAgent.agent_id.in_(body.agent_ids),
        models.SmallWorldAgent.world_id == world.id,
    ).all()

    if len(agents_db) < 2:
        raise HTTPException(status_code=400, detail="Need at least 2 agents to suggest relationships")

    agent_summaries = [
        {
            "agent_id": a.agent_id,
            "name": a.name,
            "job_title": a.job_title,
            "profession": a.profession,
            "organization": a.organization,
        }
        for a in agents_db
    ]

    try:
        suggestions, usage = suggest_relationships(agent_summaries)
    except Exception as exc:
        import traceback
        print(f"[auto-suggest-relationships] ERROR: {exc}\n{traceback.format_exc()}", flush=True)
        raise HTTPException(status_code=500, detail=f"AI suggestion failed: {exc}")

    try:
        deduct_credits(db, current_user.id, usage, description="Relationship auto-suggest")
    except Exception as billing_exc:
        print(f"[billing] deduct_credits failed: {billing_exc}")

    # Build a lookup from agent_id (UUID str) -> db id
    id_map = {a.agent_id: a for a in agents_db}

    created_rels = []
    for s in suggestions:
        src_uuid = s.get("source_agent_id")
        tgt_uuid = s.get("target_agent_id")
        if not src_uuid or not tgt_uuid or src_uuid == tgt_uuid:
            continue
        src = id_map.get(src_uuid)
        tgt = id_map.get(tgt_uuid)
        if not src or not tgt:
            continue

        # Skip if already exists
        existing = db.query(models.AgentRelationship).filter(
            models.AgentRelationship.source_agent_id == src.id,
            models.AgentRelationship.target_agent_id == tgt.id,
            models.AgentRelationship.type == s.get("type", "peer"),
        ).first()
        if existing:
            continue

        rel = models.AgentRelationship(
            source_agent_id=src.id,
            target_agent_id=tgt.id,
            type=s.get("type", "peer"),
            strength=float(s.get("strength", 0.5)),
            sentiment=s.get("sentiment", "neutral"),
            influence_direction=s.get("influence_direction", "both"),
        )
        db.add(rel)
        db.flush()
        created_rels.append((rel, src, tgt))

    db.commit()

    result = []
    for rel, src, tgt in created_rels:
        db.refresh(rel)
        result.append(schemas.AgentRelationshipResponse(
            id=rel.id,
            rel_id=rel.rel_id,
            source_agent_id=src.agent_id,
            target_agent_id=tgt.agent_id,
            source_agent_name=src.name,
            target_agent_name=tgt.name,
            type=rel.type,
            strength=rel.strength,
            sentiment=rel.sentiment,
            influence_direction=rel.influence_direction,
        ))
    return result
